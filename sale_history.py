from my_package.funcs import *
from SalePage import SalePage
import requests
from datetime import date
import openpyxl

temps = [

    '''Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7
Accept-Encoding: gzip, deflate, br
Accept-Language: zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7
Cache-Control: no-cache
Connection: keep-alive
Cookie: Device-Id=yzaXnFdnQNeqjCExrtg9; _ga=GA1.2.1712764201.1626950100; Locale-Supported=zh-Hans; game=csgo; session=1-59w3bgYYhTfBc3nmndQKyTfdDqfJtlqhyfygJ1z0501N2042474895; csrf_token=IjMzMTRlZjQ1Zjc3MzRkMjdiNzc4YmYxNzlmNmFiZmY0YmMyMzI2NDEi.FzgYDw._rgqyqVwyJ1yPJEoHCw4-fYETH0
Host: buff.163.com
Pragma: no-cache
sec-ch-ua: "Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"
sec-ch-ua-mobile: ?0
sec-ch-ua-platform: "Windows"
Sec-Fetch-Dest: document
Sec-Fetch-Mode: navigate
Sec-Fetch-Site: none
Sec-Fetch-User: ?1
Upgrade-Insecure-Requests: 1
User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36''',

'''Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7
Accept-Encoding: gzip, deflate, br
Accept-Language: zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7
Cache-Control: no-cache
Connection: keep-alive
Cookie: Device-Id=N3UFOKd1mEEIejYq9N7l; _ga=GA1.2.595925459.1636286269; Locale-Supported=zh-Hans; game=csgo; session=1-5JwcmlyzL3Dz65re7ZBYLUfEQGQTW5Q9JaKBIVye2b9h2036448739; csrf_token=ImJiMDllNzllNjY5OTY1YmEyN2VkZjM5MmJkZTYyNjQyNzc3MzQ0NWIi.FzgUxw.luVSxkkkNmwZiZihBPYgl6DEwT0
Host: buff.163.com
Pragma: no-cache
Referer: https://buff.163.com/market/sell_order/on_sale?game=csgo&mode=2,5
sec-ch-ua: "Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"
sec-ch-ua-mobile: ?0
sec-ch-ua-platform: "Windows"
Sec-Fetch-Dest: document
Sec-Fetch-Mode: navigate  
Sec-Fetch-Site: same-origin
Sec-Fetch-User: ?1
Upgrade-Insecure-Requests: 1
User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36''',

'''Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7
Accept-Encoding: gzip, deflate, br
Accept-Language: zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7
Cache-Control: max-age=0
Connection: keep-alive
Cookie: Device-Id=pjnWvM1ryqtXoR96S42T; _ga=GA1.2.1294642150.1638847524; Locale-Supported=zh-Hans; game=csgo; session=1-WXoodJciWOkMHwgaQnC4fMxT4egzdLDPUhx5Es4kN-gA2046383888; csrf_token=IjdlN2UzNjliMjgxMmFjMjFmZDE0NmRiNzkyYjE2NGIxZWMwNWViYjIi.FzgUUw.ZXHT4m2Zj9rgcrumL_xUeCHj01c
Host: buff.163.com
Referer: https://buff.163.com/market/sell_order/on_sale?game=csgo&mode=2,5
sec-ch-ua: "Google Chrome";v="111", "Not(A:Brand";v="8", "Chromium";v="111"
sec-ch-ua-mobile: ?0
sec-ch-ua-platform: "Windows"
Sec-Fetch-Dest: document
Sec-Fetch-Mode: navigate
Sec-Fetch-Site: same-origin
Sec-Fetch-User: ?1
Upgrade-Insecure-Requests: 1
User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'''
]

month = 1
begin = date(2023, month, 1)
end = date(2023, month+4, 1)
url = 'https://buff.163.com/market/sell_order/history?game=csgo&state=success&page_size=100&page_num='
sell_entries = []

session = requests.Session()
for temp in temps:
    headers = convert_headers(temp)
    print(headers)
    for i in range(1, 1000):
        print(i)
        res = session.get(url+str(i), headers=headers)
        page = SalePage(res.text)

        entries = page.getSaleEntries(begin, end)
        sell_entries.extend(entries)
        print(f'新增{len(entries)}筆')

        if page.isEmpty() or page.out_of_date:
            print('提前結束', i)
            break
    print(len(sell_entries))

file_name = f'C:\\Users\\acer\\Desktop\\國稅局資料\\{begin.month}月銷貨明細.xlsx'
try:
    wb = openpyxl.load_workbook(file_name)
    wb.close()
except FileNotFoundError:
    wb = openpyxl.Workbook()
    wb.save(file_name)

wb = openpyxl.load_workbook(file_name)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    wb.remove(sheet)

ws = wb.create_sheet(f'{month}月')

ws.append(['名稱', '金額（人民幣）', '日期', '時間'])
for sale in sell_entries:
    ws.append(sale.to_iterable())

wb.save(file_name)

